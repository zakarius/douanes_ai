import openpyxl
import pandas as pd
from vectorstore import BaseDouaneData


def read_excel_with_merged_cells(file_path, sheet_name):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    data = []

    for row in sheet.iter_rows():
        row_data = []
        for cell in row:
            if cell.value is not None:
                if cell.coordinate in sheet.merged_cells:
                    # Check if the current cell is part of a merged cell range
                    for merged_range in sheet.merged_cells.ranges:
                        if cell.coordinate in merged_range:
                            start_cell = merged_range.min_row, merged_range.min_col
                            end_cell = merged_range.max_row, merged_range.max_col
                            break
                        
                    start_col, start_row = start_cell
                    end_col, end_row = end_cell

                    # Add the value of the merged cell to the current row
                    row_data.append(sheet.cell(row=start_row, column=start_col).value)

                    # Skip the remaining cells in the merged cell range
                    for r in range(start_row, end_row + 1):
                        for c in range(start_col, end_col + 1):
                            if (r, c) != (start_row, start_col):
                                sheet.cell(row=r, column=c, value=None)
                else:
                    row_data.append(cell.value)
        data.append(row_data)
    workbook.close()
    return data


def valeur_content(row: pd.Series):
    marchandise = row["marchandise"];
    position = row['position']
    unite = row["unite"]
    ancienne_valeur = row["ancienne_valeur"]
    nouvelle_valeur = row["nouvelle_valeur"]
    return "|".join([
        str(marchandise) if marchandise else "",
        str(position) if position is not None else "",
        str(unite) if unite is not None else "",
        str(ancienne_valeur) if ancienne_valeur is not None else "",
        str(nouvelle_valeur) if nouvelle_valeur is not None else ""
    ])


def regime_content(row: pd.Series):
    regime = row["regime"]
    code = row['code']
    def_code = row["def_code"]
    def_regime = row["def_regime"]

    return "|".join([
        str(regime) if regime else "",
        str(code) if code is not None else "",
        str(def_regime) if def_regime is not None else "",
        str(def_code) if def_code is not None else "",
    ])


class Valeur(BaseDouaneData):
    BASE_COLLECTION = "valeurs"
    TEMPERATURE = 0.1

    pre_prompt = """AGENT0 est un douanier Togolais. Tu est un assistant basé sur GPT3 qui reppond de facon précise, concise et complete aux questions douanières en se basant uniquement sur tes connaissance en matière d'evaluation de la valeur en douane des marchandises, dans le style académique. Tu te base sur la fiche des valeurs actualisées en fin d'année 2022. Si la nouvelle valeur est vide, Tu utlises l'ancienne valeur. Tu evites à tout prix de se repeter dans sa réponse et ne fabrique pas de reponse si la fiche des valeur ne le permet pas.\n\n
    Marchandises | Position Tarifaire dans le Tarif Exterieur Commun de la CEDEAO| unité ou reference | ancienne valeur | nouvelle valeur
    * Pièce détachée vélo | 87141100900 8714920090 | Kg/TC X 20 / TC X40 | 700/2 500 000 / 3 500 000 DD | 700/2 500 000 / 3500 000 DD
    * Ustensiles de cuisine en série | |Carton|8000-12000|
    * TV /Sonny/Samsung/Ecran LCD 32/42/47/50 |85287200901|Unité|70 000/90 000/ 110 000/170 000 |
    * Sucre en poudre|17019910000|CST|CST|CST
    * Carreaux ordinaire|69079000000|M carré/TC X 20' |3 500 000|4 500 000
    * Chewing gum|17041000000|Kg/ TC X 20' TC X 40' |250/ DD≥1 500 000 DD≥3 000 000 |250/ DD≥2 500 000 DD≥4 000 000
    """
        

    examples = """\n\nEXEMPLES:"
    Exemple1: pièces détachées pour vélo \nReponse: Les pièces détachées pour velo classées sous les positions 87141100900 et 8714920090 sont évalués comme suit: 700F/kg, 2.500.000F/TCx20' et 3.500.000/TCx40'"
    
    Exemple2: Ustensiles de cuisine en série \nReponse: Les ustensiles de cuisine en série sont evalués en carton. La valeur par carton va de 8.000 à 12.000F"
    
    Exemple3: Télévision Samsung écran LCD \nReponse:Les télévisions de marque Samsung à ecran LCD classées sour la 85287200901 du TEC sont evaluées comme suit:\n\t- Ecran de 32 pouces : 70.000F/unité\n\t- Ecran de 42 pouces : 90.000F/unité\n\t- Ecran de 47 pouces : 110.000F/unité\n\t- Ecran de 50 pouces : 170.000F/unité"
    
    Exemple4: Sucre en poudre\nReponse: la valeur du sucre en poudre depend du Code CST (Code Spécial de Tarification) associé à sa marque et son origine"
    
    Exemple5: les carreaux ordinaire classés sous la position 69079000000 sont evalués à 4.500.000F/TCx20'. l'ancienne valeur était de 3.500.000F/TCx20'  \nReponse: La valeur des carreaux depend de leur qualité"
    
    Exemple6: Chewing gum\nReponse: les Chewing-gums classés sous la position 17041000000 du TEC peuvent être evalués par kilogramme ou par conteneur.La valeur par kilo est de 250F. le total des droits déclarés (DD) doit être au minimum de 2.500.000F pour un TCx20' et 4.000.000F pour un TCx40'. les anciennes valeurs sont respectivement de 250F, 1.500.000F et 3.000.000F
    """

    def __init__(self):
        super().__init__()
        self.data_frames_path = __file__.replace("valeur.py", "")+"data/"
        self.embeddings_path = __file__.replace("valeur.py", "")+"embeddings/"

    def df(self, data: str = BASE_COLLECTION):
        file_path = self.data_frames_path+data
        try:
            _df = pd.read_json(file_path+"_df.json", orient="records")
        except:
            excel_data = pd.read_excel(open(file_path+".xlsx", "rb"), header=0,).values;
            _df = pd.DataFrame(columns=["marchandise", "position", "unite" ,"ancienne_valeur", "nouvelle_valeur"], data=excel_data)
            _df = _df[~(_df[['position', 'unite', 'ancienne_valeur', 'nouvelle_valeur']].isna().all(axis=1))]
            _df["position"] = _df.position.apply(lambda x: None if (str(x) == "nan" )else str(x).replace(".", "").replace(" ", ""))
            _df.fillna(" ", inplace=True)
            _df["content"] = _df.apply(valeur_content, axis=1)
            _df = _df[['content']]
            _df.to_json(file_path+"_df.json", orient="records")
        _df.reset_index(inplace=True)
        return _df

    def ge_cst(self, question: str,
               show_prompt: bool = False,
               prompt_only: bool = False,
               n_result: int = 5,
               stream: bool = False,):
        return self.answer(question, show_prompt, prompt_only, n_result, stream)


class RegimeEconomique(Valeur):
    BASE_COLLECTION = "regimes"
    TEMPERATURE = 0

    pre_prompt = """AGENT0 est un douanier Togolais. Tu est un assistant basé sur GPT3 qui reppond de facon précise, concise et complete aux questions douanières en se basant uniquement sur tes connaissance en matière régimes économiques en douanes des marchandises, dans le style académique. Tu te base sur la liste des régimes et codes additionnels contenus dans SYDONIA++. Tu evites à tout prix de se repeter dans sa réponse, ne fabrique pas de reponse si la liste des régimes ne le permet pas et precise les codes additionnels (si necessaire).\n\n
    Régime etendu | Code additionnel | libéllés régime etendu | Libellés Code additionnel 
    """

    examples = ""

    def df(self, data: str = BASE_COLLECTION):
        file_path = self.data_frames_path+data
        try:
            _df = pd.read_json(file_path+"_df.json", orient="records")
        except:
            excel_data = pd.read_excel(
                open(file_path+".xlsx", "rb"), header=0, dtype=str).values
            _df = pd.DataFrame(columns=["regime", "code", "def_regime",
                               "def_code"], data=excel_data, dtype=str)
            
            _df["content"] = _df.apply(regime_content, axis=1)
            _df = _df[['content']]
            _df.to_json(file_path+"_df.json", orient="records")
        _df.reset_index(inplace=True)
        return _df

    def get_info(self,
               question: str,
               show_prompt: bool = False,
               prompt_only: bool = False,
               n_result: int = 10,
               stream: bool = False,
               ):
        question = question.strip()
        if(len(question.split(" ")) == 1):
            if(len(question) > 4):
                question = "Régime" +  question[:4] + " "+ question[4:]

        return self.answer(question, show_prompt, prompt_only, n_result, stream)